from __future__ import annotations

import argparse
import json
import math
import statistics
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_INPUT = ROOT / "results/Resultados_Evaluacion.xlsx"


# Lectura del Excel
def read_rows(workbook, sheet_name: str) -> list[dict]:
    sheet = workbook[sheet_name]
    headers = [cell.value for cell in sheet[3]]
    return [dict(zip(headers, values)) for values in sheet.iter_rows(min_row=4, values_only=True)]


# Cálculos básicos
def mean(values) -> float:
    data = list(values)
    return statistics.fmean(data) if data else 0.0


def percentile(values, q: float) -> float:
    data = sorted(values)
    position = (len(data) - 1) * q
    lower, upper = math.floor(position), math.ceil(position)
    if lower == upper:
        return data[lower]
    return data[lower] + (data[upper] - data[lower]) * (position - lower)


def retrieval(rows: list[dict], column: str) -> dict:
    ranks = [row[column] for row in rows]
    return {
        "recall_at_1": mean(rank is not None and rank <= 1 for rank in ranks),
        "recall_at_3": mean(rank is not None and rank <= 3 for rank in ranks),
        "recall_at_5": mean(rank is not None and rank <= 5 for rank in ranks),
        "recall_at_10": mean(rank is not None and rank <= 10 for rank in ranks),
        "mrr": mean(1 / rank if rank else 0 for rank in ranks),
    }


# Evaluación final
def evaluate(path: Path) -> dict:
    workbook = load_workbook(path, read_only=True, data_only=True)
    t1 = read_rows(workbook, "T1 IoC")
    t2 = read_rows(workbook, "T2 CVE")
    t3 = read_rows(workbook, "T3 ATTACK")
    t4 = read_rows(workbook, "T4 CTI")
    latency = read_rows(workbook, "Latencia")

    tp = sum(row["TP"] for row in t1)
    fp = sum(row["FP"] for row in t1)
    fn = sum(row["FN"] for row in t1)
    precision = tp / (tp + fp)
    recall = tp / (tp + fn)
    f1 = 2 * precision * recall / (precision + recall)

    result = {
        "T1_IOC": {
            "precision_micro": precision,
            "recall_micro": recall,
            "f1_micro": f1,
        },
        "T2_CVE": {},
        "T3_ATTACK": {},
        "T4_CTI": {},
        "latency": {},
    }

    for condition in ("B0", "R1"):
        rows = [row for row in t2 if row["Condición"] == condition]
        result["T2_CVE"][condition] = {
            "raw_accuracy": mean(row["Respuesta correcta"] for row in rows),
            "grounded_accuracy": mean(row["Correcta y fundamentada"] for row in rows),
            "abstention_rate": mean(row["Abstención"] for row in rows),
        }

        rows = [row for row in t3 if row["Condición"] == condition]
        result["T3_ATTACK"][condition] = {
            "exact_top1_accuracy": mean(row["Exactitud top-1"] for row in rows),
            "abstention_rate": mean(row["Abstención"] for row in rows),
        }
        if condition == "R1":
            result["T3_ATTACK"][condition]["retrieval"] = retrieval(rows, "Posición recuperada")

        rows = [row for row in t4 if row["Condición"] == condition]
        result["T4_CTI"][condition] = {
            "automatic_grounded_rate": mean(row["Fundamentación automática"] for row in rows),
            "abstention_rate": mean(row["Abstención"] for row in rows),
            "factual_correctness_mean_0_2": mean(row["Corrección factual"] for row in rows),
            "completeness_mean_0_2": mean(row["Completitud"] for row in rows),
            "unsupported_claims_mean_0_2": mean(row["No respaldadas"] for row in rows),
        }
        if condition == "R1":
            result["T4_CTI"][condition]["document_retrieval"] = retrieval(rows, "Posición documento")
            result["T4_CTI"][condition]["passage_retrieval"] = retrieval(rows, "Posición pasaje")

    for condition in ("DIRECT", "B0", "R1"):
        values = [row["Extremo a extremo (s)"] for row in latency if row["Condición"] == condition]
        result["latency"][condition] = {
            "requests": len(values),
            "median_seconds": statistics.median(values),
            "p95_seconds": percentile(values, 0.95),
        }
    return result


# Argumentos y salida
def main() -> int:
    parser = argparse.ArgumentParser(description="Recalcula las métricas principales de la evaluación")
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--output", type=Path, help="Archivo JSON opcional para guardar las métricas")
    args = parser.parse_args()

    result = evaluate(args.input)
    text = json.dumps(result, ensure_ascii=False, indent=2)
    print(text)
    if args.output:
        args.output.parent.mkdir(parents=True, exist_ok=True)
        args.output.write_text(text + "\n", encoding="utf-8")
        print(f"Métricas guardadas en {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
